"""
Reorganiza o export do TARGIT de Meta de Tonelada (formato "largo",
uma coluna por RCA/mês) num CSV normal (uma linha por
filial/ano/mês/RCA), pronto pra ser usado pelo chatbot.

Uso:
    python3 scripts/reorganizar_meta_tonelada.py caminho/do/Export.xlsx

Gera dados/meta_tonelada_2024_2026.csv.
"""
import sys
from pathlib import Path

import openpyxl
import pandas as pd

MAPA_MES = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8,
    "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}

RAIZ_PROJETO = Path(__file__).resolve().parent.parent
ARQUIVO_SAIDA = RAIZ_PROJETO / "dados" / "meta_tonelada_2024_2026.csv"


def reorganizar(caminho_entrada: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(caminho_entrada, data_only=True)
    ws = wb[wb.sheetnames[0]]

    linha_rca = [c.value for c in ws[1]]
    linha_mes = [c.value for c in ws[2]]
    linha_medida = [c.value for c in ws[3]]

    registros = []

    for row in ws.iter_rows(min_row=4, values_only=True):
        filial = row[0]
        ano = row[1]

        if filial is None or ano is None:
            continue
        if ano == "Total" or filial == "Total":
            continue

        for col_idx in range(2, len(row)):
            valor = row[col_idx]
            if valor is None:
                continue

            rca = linha_rca[col_idx]
            mes = linha_mes[col_idx]
            medida = linha_medida[col_idx]

            if rca is None or mes is None or medida is None:
                continue
            if mes == "Total":
                continue

            registros.append({
                "FILIAL": filial,
                "ANO": int(ano),
                "MES": MAPA_MES.get(mes),
                "RCA": rca,
                "MEDIDA": medida,
                "VALOR": float(valor),
            })

    df_longo = pd.DataFrame(registros)

    df_final = (
        df_longo
        .pivot_table(
            index=["FILIAL", "ANO", "MES", "RCA"],
            columns="MEDIDA",
            values="VALOR",
            aggfunc="first",
        )
        .reset_index()
    )
    df_final.columns.name = None

    return df_final


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python3 scripts/reorganizar_meta_tonelada.py <export.xlsx>")
        sys.exit(1)

    df = reorganizar(sys.argv[1])
    ARQUIVO_SAIDA.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(ARQUIVO_SAIDA, sep=";", index=False)

    print(f"Linhas geradas: {len(df)}")
    print(f"Salvo em: {ARQUIVO_SAIDA}")
